
import os
import sys
import openpyxl
from datetime import datetime

# Add parent directory to path to import app modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import app
from extensions import db
from models import Equipment, User
from config import Config

def migrate_data():
    print("Iniciando migración de datos (Motor: OpenPyXL)...")
    
    with app.app_context():
        # Crear tablas
        db.create_all()
        
        if Equipment.query.first():
            print("¡La base de datos ya contiene datos! Abortando.")
            return

        try:
            excel_path = Config.EXCEL_PATH_LEGACY
            print(f"Leyendo Excel desde: {excel_path}")
            
            # Cargar workbook
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = wb.active
            
            # Auto-detect header row
            header_row_idx = 1
            headers = []
            
            # Buscar fila de encabezados en las primeras 10 filas
            for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
                temp_headers = [str(h).strip().upper() if h else '' for h in row]
                if 'FR' in temp_headers and 'MARCA' in temp_headers:
                    headers = temp_headers
                    print(f"Encabezados encontrados en fila {header_row_idx}: {headers}")
                    break
                header_row_idx += 1
            
            if not headers:
                print("No se encontró la fila de encabezados (buscando FR y MARCA)")
                return

            # Mapeo de indices
            try:
                idx_fr = headers.index('FR')
                idx_marca = headers.index('MARCA')
                idx_modelo = headers.index('MODELO')
                idx_reporte = headers.index('REPORTE DE CLIENTE')
                idx_estado = headers.index('ESTADO')
                idx_condicion = headers.index('CONDICION') if 'CONDICION' in headers else -1
                idx_encargado = headers.index('ENCARGADO') if 'ENCARGADO' in headers else -1
            except ValueError as e:
                print(f"Error: Columna faltante en el Excel: {e}")
                return

            migrated_count = 0
            
            # Iterar filas (empezando después de la fila de encabezados)
            for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                # Si la fila está vacía o no tiene FR, saltar
                if not row[idx_fr]:
                    continue
                    
                try:
                    equipment = Equipment(
                        fr=str(row[idx_fr]),
                        marca=str(row[idx_marca]) if row[idx_marca] else None,
                        modelo=str(row[idx_modelo]) if row[idx_modelo] else None,
                        reporte_cliente=str(row[idx_reporte]) if row[idx_reporte] else None,
                        estado=str(row[idx_estado]).strip().upper() if row[idx_estado] else 'PENDIENTE',
                        condicion=str(row[idx_condicion]) if idx_condicion != -1 and row[idx_condicion] else None,
                        encargado=str(row[idx_encargado]) if idx_encargado != -1 and row[idx_encargado] else None,
                        fecha_ingreso=datetime.utcnow()
                    )
                    
                    db.session.add(equipment)
                    migrated_count += 1
                except Exception as e:
                    print(f"Error importando fila: {e}")

            # Crear Admin
            if not User.query.filter_by(username='admin').first():
                admin = User(username='admin', is_admin=True, is_approved=True)
                admin.set_password('admin123')
                db.session.add(admin)
                print("Usuario 'admin' creado.")

            db.session.commit()
            print(f"Migración completada. Registros: {migrated_count}")

        except Exception as e:
            print(f"Error fatal: {e}")

if __name__ == "__main__":
    migrate_data()
